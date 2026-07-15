import re
import time
import random
import logging
from dataclasses import dataclass, field, asdict
from typing import Optional

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

SEARCH_KEYWORD = "wireless mouse"
PAGES_TO_SCRAPE = 2          # first two pages as requested
BASE_SEARCH_URL = "https://www.ebay.com/sch/i.html"
OUTPUT_FILE = "ebay_wireless_mouse_results.xlsx"
REQUEST_TIMEOUT = 15
MIN_DELAY = 1.5              # polite delay range between requests (seconds)
MAX_DELAY = 3.5
MAX_ITEMS = None             # set an int (e.g. 40) to cap items for testing

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("ebay_scraper")

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
    "Sec-Fetch-User": "?1",
}

MAX_RETRIES = 3
RETRY_BACKOFF_BASE = 3  # seconds; doubles each retry


def parse_html(html_text: str) -> BeautifulSoup:
    """Parse HTML, falling back to the stdlib parser if lxml isn't installed."""
    try:
        return BeautifulSoup(html_text, "lxml")
    except Exception:
        return BeautifulSoup(html_text, "html.parser")


def warm_up_session(session: requests.Session):
    """Hit the eBay homepage first so the session picks up normal cookies
    before we start requesting search/item pages. Reduces first-request 403s."""
    try:
        session.get("https://www.ebay.com/", headers=HEADERS, timeout=REQUEST_TIMEOUT)
        time.sleep(random.uniform(1, 2))
    except requests.RequestException as exc:
        log.warning("Warm-up request failed (continuing anyway): %s", exc)

ITEM_URL_RE = re.compile(r"https://www\.ebay\.com/itm/\d+[^\"'\s]*")


@dataclass
class Listing:
    title: Optional[str] = None
    price: Optional[str] = None
    condition: Optional[str] = None
    availability: Optional[str] = None
    sold: Optional[str] = None
    image_url: Optional[str] = None
    seller_name: Optional[str] = None
    seller_feedback: Optional[str] = None
    item_url: str = ""
    error: Optional[str] = None


def polite_sleep():
    time.sleep(random.uniform(MIN_DELAY, MAX_DELAY))


def get_soup(session: requests.Session, url: str, params: dict = None) -> Optional[BeautifulSoup]:
    """Fetch a URL and return a parsed BeautifulSoup object, or None on failure.
    Retries on 403/429/5xx with exponential backoff before giving up."""
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = session.get(url, headers=HEADERS, params=params, timeout=REQUEST_TIMEOUT)
        except requests.RequestException as exc:
            log.warning("Request error for %s (attempt %d/%d): %s", url, attempt, MAX_RETRIES, exc)
            resp = None

        if resp is not None and resp.status_code == 200:
            return parse_html(resp.text)

        status = resp.status_code if resp is not None else "no response"
        if attempt < MAX_RETRIES:
            wait = RETRY_BACKOFF_BASE * (2 ** (attempt - 1))
            log.warning("Status %s for %s - retrying in %ds (attempt %d/%d)",
                        status, url, wait, attempt, MAX_RETRIES)
            time.sleep(wait)
        else:
            log.warning("Giving up on %s after %d attempts (last status: %s)",
                        url, MAX_RETRIES, status)

    return None


def collect_item_links(session: requests.Session) -> list:
    """Go through PAGES_TO_SCRAPE search-result pages and collect unique item URLs."""
    links = []
    seen = set()

    for page_num in range(1, PAGES_TO_SCRAPE + 1):
        params = {
            "_nkw": SEARCH_KEYWORD,
            "_sacat": 0,
            "_pgn": page_num,
        }
        log.info("Fetching search results page %d ...", page_num)
        soup = get_soup(session, BASE_SEARCH_URL, params=params)
        if soup is None:
            log.warning("Search page %d could not be loaded after retries - skipping", page_num)
            continue

        # Item links are matched by regex directly against the raw HTML.
        # This is more robust than chasing eBay's frequently-changing CSS
        # class names on the search grid.
        found = ITEM_URL_RE.findall(str(soup))

        page_new = 0
        for url in found:
            clean_url = url.split("?")[0]  # strip tracking params, keep canonical /itm/<id>
            if clean_url not in seen:
                seen.add(clean_url)
                links.append(clean_url)
                page_new += 1

        log.info("Page %d: found %d new unique item links (running total: %d)",
                  page_num, page_new, len(links))
        polite_sleep()

    return links


def text_or_none(el) -> Optional[str]:
    return el.get_text(strip=True) if el else None


def scrape_item(session: requests.Session, url: str) -> Listing:
    listing = Listing(item_url=url)

    soup = get_soup(session, url)
    if soup is None:
        listing.error = "Failed to load page"
        return listing

    # ---- Title -----------------------------------------------------
    title_el = soup.select_one("h1.x-item-title__mainTitle span.ux-textspans")
    listing.title = text_or_none(title_el)

    # ---- Price -------------------------------------------------------
    price_el = soup.select_one("div.x-price-primary span.ux-textspans")
    listing.price = text_or_none(price_el)

    # ---- Condition -----------------------------------------------------
    cond_el = soup.select_one("div.x-item-condition-text span.ux-textspans") \
        or soup.select_one("div.x-item-condition-text")
    listing.condition = text_or_none(cond_el)

    # ---- Availability & Sold -------------------------------------------
    # Both live as sibling <span class="ux-textspans ux-textspans--SECONDARY">
    # inside div#qtyAvailability. First is availability, second (if present) is sold count.
    qty_block = soup.select_one("div#qtyAvailability")
    if qty_block:
        secondary_spans = qty_block.select("span.ux-textspans--SECONDARY")
        texts = [s.get_text(strip=True) for s in secondary_spans if s.get_text(strip=True)]
        if texts:
            listing.availability = texts[0]
        if len(texts) > 1:
            listing.sold = texts[1]
        # Fallback: sometimes "available" and "sold" render as separate text nodes
        if listing.sold is None:
            full_text = qty_block.get_text(" ", strip=True)
            m = re.search(r"([\d,]+)\s+sold", full_text)
            if m:
                listing.sold = f"{m.group(1)} sold"

    # ---- Image -----------------------------------------------------
    img_el = soup.select_one("img.img-scale-down") or soup.select_one("div.ux-image-carousel-item.active img")
    if img_el:
        listing.image_url = img_el.get("src") or img_el.get("data-zoom-src")

    # ---- Seller -----------------------------------------------------
    seller_el = soup.select_one("div[data-testid='x-sellercard-atf'] a span.ux-textspans--BOLD") \
        or soup.select_one("div[data-testid='x-sellercard-atf'] a")
    listing.seller_name = text_or_none(seller_el)

    feedback_el = soup.select_one("div[data-testid='x-sellercard-atf'] span.ux-textspans--PSEUDOLINK") \
        or soup.select_one("div[data-testid='x-sellercard-atf'] a[href*='fdbk']")
    listing.seller_feedback = text_or_none(feedback_el)

    return listing


def build_excel(listings: list, filepath: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "eBay Wireless Mouse"

    headers = [
        "Title", "Price", "Condition", "Availability", "Sold",
        "Image URL", "Seller", "Seller Feedback", "Item URL", "Notes",
    ]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for item in listings:
        ws.append([
            item.title, item.price, item.condition, item.availability,
            item.sold, item.image_url, item.seller_name,
            item.seller_feedback, item.item_url, item.error,
        ])

    # rough auto-width
    widths = [40, 14, 20, 20, 12, 55, 18, 16, 55, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    wb.save(filepath)
    log.info("Saved %d rows to %s", len(listings), filepath)


def main():
    session = requests.Session()

    log.info("Warming up session ...")
    warm_up_session(session)

    log.info("Stage 1: collecting item links from %d search page(s) for '%s'",
              PAGES_TO_SCRAPE, SEARCH_KEYWORD)
    item_links = collect_item_links(session)

    if not item_links:
        log.error("No item links were found. eBay may have blocked the request "
                   "(try again later, or reduce request speed / add a rotating "
                   "User-Agent / use a proxy).")
        return

    if MAX_ITEMS:
        item_links = item_links[:MAX_ITEMS]

    log.info("Stage 2: visiting %d item pages for full details ...", len(item_links))

    results = []
    for i, url in enumerate(item_links, start=1):
        log.info("[%d/%d] %s", i, len(item_links), url)
        listing = scrape_item(session, url)
        results.append(listing)
        polite_sleep()

    build_excel(results, OUTPUT_FILE)

    ok = sum(1 for r in results if not r.error)
    log.info("Done. %d/%d items scraped successfully.", ok, len(results))


if __name__ == "__main__":
    main()